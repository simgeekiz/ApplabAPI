from openpyxl import load_workbook
import numpy as np
import datetime
from copy import deepcopy
import pandas as pd

import os
import sys
import json

sys.path.append('..')




# helper function
def get_date_from_str(yyyymmdd):
    """Return a datetime representation of the date

    Parameters
    ----------
    yyyymmdd : string,
        Date string in format yyyy-mm-dd. Division symbols can differ

    Returns
    -------
    date : datetime,
        A datetime representation of the date
    """
    split_token = yyyymmdd[4]
    year, month, day = [int(s) for s in yyyymmdd.split(split_token)]
    #print("{}:{}:{}".format(year, month, day))
    return datetime.datetime(year, month, day, 0, 0, 0, 0)

class MyEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        else:
            return super(MyEncoder, self).default(obj)


class DataHandler():

    def __init__(self, submissions=None):
        df = pd.DataFrame(submissions)
        self.dates = np.array(df['SubmitDateTime'])
        self.user_ids = np.array(df['UserId'])
        self.exercise_ids = np.array(df['ExerciseId'])
        self.learn_obj_ids = np.array(df['LearningObjectiveId'])
        self.corrects = np.array(df['Correct'])
        self.ability_scores = np.array(df['AbilityAfterAnswer'])
        self.m2m = MomentByMoment(self.user_ids, self.corrects, self)
        # release memory
        self.avg_ltgs = [[0.999, 0.125, 0.299, 0.099],
                        [0.5, 0.1, 0.299, 0.1],
                        [0.5, 0.1, 0.299, 0.1],
                        [0.5, 0.1, 0.299, 0.1],
                        ]
        del df

    def get_max_row(self, column=1):
        val = self.ws.cell(row=1, column=column).value
        i = 1
        while val:
            i += 1
            val = self.ws.cell(row=i, column=column).value
        return i

    def get_color_row(self, column=1):
        for i in range(1, 20):
            color = self.ws.cell(row=i, column=column).fill.start_color.index
            if color != '00000000':
                return i

    def get_column(self, cid, start_row=2, end_row=None):
        if end_row is None:
            end_row = self.max_row
        if end_row <= start_row:
            end_row = start_row + 1
        column = []
        for i in range(start_row, end_row):
            value = self.ws.cell(row=i, column=cid).value
            if value or value == 0:
                column.append(value)
        return np.array(column)

    def get_users(self):
        return np.unique(self.user_ids)

    def get_corrects_from_user(self, user_id=0, method='all'):
        if method == 'all':
            return self.corrects[np.where(self.user_ids == user_id)]

    def get_dummy_graph_variables(self, user_id=0, method='all', oid=None):
        return self.m2m.get_p_T(user_id, method, oid)

    def get_graph_variables(self, user_id=0, method='all', oid=None):
        p_j = self.m2m.get_p_j(user_id=user_id, method=method,
                               objective_id=oid)
        self.graph_length = len(p_j)
        self.boundary_list, self.color_list = self.m2m.get_color_bars()
        return self.m2m.get_p_j(user_id=user_id, method=method,
                                objective_id=oid)

    def get_color_ids(self, fname='res/ID exercises.xlsx'):
        wb = load_workbook(fname)
        self.ws = wb.active
        pre = list(self.get_column(1, 3))
        c_bound = [self.get_color_row(i) for i in range(2, 5)]
        cin = list(self.get_column(2, 3, c_bound[0])) + \
              list(self.get_column(3, 3, c_bound[0])) + \
              list(self.get_column(4, 3, c_bound[0]))
        cex = list(self.get_column(2, c_bound[0])) + \
              list(self.get_column(3, c_bound[0])) + \
              list(self.get_column(4, c_bound[0]))
        aex = None
        raex = None
        post = list(self.get_column(5, 3))
        return (pre, cin, cex, aex, raex, post)

    def get_coordinates_for_today(self, user_id=0, loid=0,
                                  date_id=None):
        # todo: change from answers to coordinates
        if date_id is None:
            answers = self.corrects[np.where((self.user_ids == user_id)&
                                             (self.learn_obj_ids == loid))]
        else:
            datelist = []
            for sdt in self.dates:
                try:
                    datelist.append(datetime.datetime.strptime(sdt, "%Y-%m-%d %H:%M:%S.%f").replace(hour=0, minute=0, second=0, microsecond=0))
                except ValueError:
                    datelist.append(datetime.datetime.strptime(sdt, "%Y-%m-%d %H:%M:%S").replace(hour=0, minute=0, second=0, microsecond=0))
                except TypeError:
                    datelist.append(None)
            datelist = np.array(datelist)
            answers = self.corrects[np.where((self.user_ids == user_id) &
                                             (datelist == date_id) &
                                             # date_id is 2017-11-28 but self dates= 2017-11-28 09:49:11.913
                                             (self.learn_obj_ids == loid))]
            #if (datelist == date_id):
            #   print("entered id")
            print("date_id",type(date_id), type(datelist[0]), date_id==datelist[0])
        same = list(np.ones(len(answers)))
        try:
            same[0] = 0
        except IndexError:
            return []
        try:
            with open("../../data/parameters.csv") as f:
                parameters = np.genfromtxt(f)
                users = parameters[:1]
                parameters = parameters[np.where(str(user_id) == users)]
                parameters = parameters[1:]
        except FileNotFoundError:
            parameters = ParameterExtractor().smart_ssr(answers, same, 100, 3)
        user_m2m = deepcopy(self.m2m)
        user_m2m.set_initial_probabilities(*parameters)
        p_j = user_m2m.get_p_j(user_id, answers=answers)
        coordinates = self.from_p_j_to_coord(p_j)
        return coordinates

    def from_p_j_to_coord(self, p_j):
        summed_array = np.cumsum(p_j)[::-1]
        #print(summed_array)
        direction = 1
        coords = []
        current = 0
        for speed in summed_array:
            current += direction * speed
            if abs(current) > speed:
                current = min(speed, max(-1 * speed, current))
                direction *= -1
                current += direction * speed
            coords.append(round(current, 2))
        try:
            max_val = coords[np.argmax(coords)]
        except Exception as e:
            return [0.0]
        if max_val > 0:
            coords = [i/max_val for i in coords]
        return coords

    def fast_coords_for_today(self, user_id, corrects, day):
        parameters = self.avg_ltgs[day]
        user_m2m = deepcopy(self.m2m)
        user_m2m.set_initial_probabilities(*parameters)
        p_j = user_m2m.get_p_j(user_id, answers=corrects)
        coordinates = self.from_p_j_to_coord(p_j)
        return coordinates


class MomentByMoment():

    def __init__(self, user_ids, corrects, handler: DataHandler):
        self.p_l0 = 0.064
        self.p_T = 0.095
        self.p_G = 0.299
        self.p_S = 0.1
        self.p_ln = []
        self.user_ids = user_ids
        self.users = np.unique(user_ids)
        self.answers = corrects
        self.handler = handler
        self.list_p_T = []

    def set_initial_probabilities(self, l0, T, G, S):
        self.p_l0 = l0
        self.p_T = T
        self.p_G = G
        self.p_S = S

    def get_p_T(self, user_id, method='all', objective_id=None):
        return self.list_p_T[:]

    def get_p_j(self, user_id, method='all', objective_id=None, answers=None):
        user_answers = self.filter_answers(user_id, method, objective_id)
        if answers is not None:
            user_answers = answers
        p_ln = self.calculate_ln(user_answers)
        p_not_ln_t = [(1 - ln) * self.p_T for ln in p_ln]
        p_not_ln_not_t = [(1 - ln) * (1 - self.p_T) for ln in p_ln]
        self.list_p_T = p_ln[:-2]
        return self.calculate_p_j(user_answers, p_ln, p_not_ln_t,
                                  p_not_ln_not_t)

    def filter_answers(self, user_id, method, objectives_id):
        user_answers = self.answers[:]
        user_objectives = self.handler.learn_obj_ids[:]
        user_ids = self.user_ids[:]
        user_excs = self.handler.exercise_ids[:]
        user_dates = self.handler.dates[:]
        self.chosen_ids = np.where(
            (self.handler.learn_obj_ids == objectives_id) &
            (self.user_ids == user_id))

        user_answers = user_answers[self.chosen_ids]
        user_objectives = user_objectives[self.chosen_ids]
        user_excs = user_excs[self.chosen_ids]
        user_ids = user_ids[self.chosen_ids]
        if method not in ['all', 'first', 'second', 'all but first', 'last']:
            raise NotImplementedError
        if method == 'first':
            user_answers = self.filter_all_but_first(user_answers, user_excs)
        if method == 'second':
            user_answers = self.filter_all_but_second(user_answers, user_excs)
        if method == 'all but first':
            user_answers = self.filter_first(user_answers, user_excs)
        if method == 'last':
            user_answers = self.filter_all_but_last(user_answers, user_excs)
        self.excs = user_excs
        return user_answers

    def filter_all_but_first(self, answers, exercise_ids):
        exercises_processed = []
        return_answers = []
        for i in range(len(answers)):
            if exercise_ids[i] not in exercises_processed:
                return_answers.append(answers[i])
                exercises_processed.append(exercise_ids[i])
        return return_answers

    def filter_all_but_second(self, answers, exercise_ids):
        exercises_processed = []
        exercises_processed_twice = []
        return_answers = []
        for i in range(len(answers)):
            if exercise_ids[i] not in exercises_processed:
                exercises_processed.append(exercise_ids[i])
            else:
                if exercise_ids[i] not in exercises_processed_twice:
                    return_answers.append(answers[i])
                    exercises_processed_twice.append(exercise_ids[i])
        return return_answers

    def filter_first(self, answers, exercise_ids):
        exercises_processed = []
        return_answers = []
        for i in range(len(answers)):
            if exercise_ids[i] not in exercises_processed:
                exercises_processed.append(exercise_ids[i])
            else:
                return_answers.append(answers[i])
        return return_answers

    def filter_all_but_last(self, answers, exercise_ids):
        answers = answers[::-1]
        exercises_processed = []
        return_answers = []
        for i in range(len(answers)):
            if exercise_ids[i] not in exercises_processed:
                return_answers.append(answers[i])
                exercises_processed.append(exercise_ids[i])
        return return_answers[::-1]

    def calculate_ln(self, answers):
        p_ln = []
        for answer_id in range(len(answers)):
            if len(p_ln) == 0:
                k = self.p_l0
            else:
                k = p_ln[-1]
            s = self.p_S
            g = self.p_G
            if answers[answer_id] == 1:
                ln_prev_given_res = (k * (1 - s)) / (
                (k * (1 - s)) + ((1 - k) * g))
            else:
                ln_prev_given_res = (k * s) / ((k * s) + ((1 - k) * (1 - g)))
            p_ln.append(ln_prev_given_res + (1 - ln_prev_given_res) * self.p_T)
        return p_ln

    def calculate_p_j(self, answers, ln, n_ln_t, n_ln_n_t):
        p_j = []
        for a_id in range(len(answers) - 2):
            p_l = ln[a_id]
            p_nl_t = n_ln_t[a_id]
            x = n_ln_n_t[a_id]
            g = self.p_G
            t = self.p_T
            s = self.p_S
            if answers[a_id + 1] == 1:
                if answers[a_id + 2] == 1:  # RR
                    a_ln = (1 - s) ** 2
                    a_n_ln_n_t = g * (1 - t) * g + g * t * (1 - s)
                else:  # RW
                    a_ln = s * (1 - self.p_S)
                    a_n_ln_n_t = g * (1 - t) * (1 - g) + g * t * s
            else:
                if answers[a_id + 2] == 1:  # WR
                    a_ln = s * (1 - self.p_S)
                    a_n_ln_n_t = g * (1 - t) * (1 - g) + (1 - g) * t * (1 - s)
                else:  # WW
                    a_ln = s ** 2
                    a_n_ln_n_t = (1 - g) * (1 - t) * (1 - g) + (1 - g) * t * s
            a_n_ln_t = a_ln
            a12 = p_l * a_ln + p_nl_t * a_n_ln_t + x * a_n_ln_n_t
            p_j.append(a_n_ln_t * p_nl_t / a12)
        return p_j

    def get_color_bars(self):
        excs = self.excs
        bounds = [0]
        colors = ['blue', 'red', 'green', 'orange', 'purple', 'magenta']
        # find pre bound
        # print('finding pre-test')
        bound = 0
        e = excs[0]
        #print(e)
        while e in self.handler.pre_ids:
            bound += 1
            e = excs[bound]
        # print(e)
        bounds.append(bound)

        # Find class instruction
        # print('finding instruction exercises')
        bound += 1
        e = excs[bound + 1]
        while e in self.handler.c_in_ids:
            bound += 1
            e = excs[bound]
        # print(e)
        bounds.append(bound)

        # Find class exercise
        # print('finding class exercises')
        bound += 1
        e = excs[bound + 1]
        while e in self.handler.c_ex_ids:
            bound += 1
            e = excs[bound]
        # print(e)
        bounds.append(bound)

        # Find class adaptive
        bound += 1
        dates = self.handler.dates[self.chosen_ids]
        d = dates[bound]
        while dates[bound + 1] == d:
            bound += 1
        # print(excs[bound])
        bounds.append(bound)

        # Find repetition adaptive
        # print('finding repeated adaptive exercises')
        bound += 1
        e = excs[bound + 1]
        while e not in self.handler.post_ids:
            bound += 1
            e = excs[bound]
        # print(e)
        bounds.append(bound)

        # print('find post-test')
        while bound < len(excs) - 1:
            bound += 1
        # print(excs[bound])
        bounds.append(bound)
        # print(bounds)
        return (bounds, colors)


class ParameterExtractor():

    def __init__(self):
        # params = [L0, T, G, S]
        self.params_min = [1e-15 for i in range(4)]
        self.params_max = [1.0, 1.0, 0.3, 0.1]

    def brute_force_params(self, answers, same, grain=100, L0_fix=None,
                           T_fix=None, G_fix=None, S_fix=None, ):
        # set ranges up
        best_l0 = L0_fix
        best_t = T_fix
        best_g = G_fix
        best_s = S_fix
        best_SSR = len(answers) * 999999999999991
        L0_range = self.get_range(L0_fix, 0, grain)
        T_range = self.get_range(T_fix, 1, grain)
        G_range = self.get_range(G_fix, 2, grain)
        S_range = self.get_range(S_fix, 3, grain)
        for L0 in L0_range:
            # print('------------------------------------\nL0 is now at:{}'.format(L0))
            for T in T_range:
                for G in G_range:
                    for S in S_range:
                        new_SSR = self.get_s_s_r(L0, T, G, S, answers, same)
                        if new_SSR < best_SSR:
                            best_l0, best_t, best_g, best_s, best_SSR = [L0, T,
                                                                         G, S,
                                                                         new_SSR]
                        # print('best parameters now at L0:{}, T:{}, G:{}, S:{}'.format(L0,
                        # 	T, G, S))
        return best_l0, best_t, best_g, best_s

    def get_s_s_r(self, L0, T, G, S, answers, sames=None):
        SSR = 0.0
        S = max(1E-15, S)
        T = max(1E-15, T)
        G = max(1E-15, G)
        L0 = max(1E-15, L0)
        L = L0
        for same, answer in zip(sames, answers):
            if same == 0:
                L = L0
            # print(L, T, G, S)
            SSR += (answer - (L * (1.0 - S) + (1.0 - L) * G)) ** 2
            if answer == 0:
                L_given_answer = (L * S) / ((L * S) + ((1.0 - L) * (1.0 - G)))
            else:
                L_given_answer = (L * (1.0 - S)) / (
                (L * (1.0 - S)) + ((1.0 - L) * G))
            if not L_given_answer:
                print('huh')
            L = L_given_answer + (1.0 - L_given_answer) * T
        return SSR

    def get_range(self, possible_range, par_id, grain):
        if possible_range is None:
            return np.linspace(self.params_min[par_id],
                               self.params_max[par_id],
                               int(grain * self.params_max[par_id]),
                               endpoint=False)[1:]
        return [possible_range]

    def smart_ssr(self, answers, same, grain, iterations):
        best_l0 = \
        self.brute_force_params(answers, same, grain, None, 0.0, 0.0, 0.0)[0]
        best_t = \
        self.brute_force_params(answers, same, grain, 0.0, None, 0.0, 0.0)[1]
        best_g = \
        self.brute_force_params(answers, same, grain, 0.0, 0.0, None, 0.0)[2]
        best_s = \
        self.brute_force_params(answers, same, grain, 0.0, 0.0, 0.0, None)[3]
        for i in range(iterations):
            # print("best is {}".format([best_l0, best_t, best_g, best_s]))
            best_l0 = \
            self.brute_force_params(answers, same, grain, None, best_t, best_g,
                                    best_s)[0]
            best_t = \
            self.brute_force_params(answers, same, grain, best_l0, None,
                                    best_g,
                                    best_s)[1]
            best_g = \
            self.brute_force_params(answers, same, grain, best_l0, best_t,
                                    None,
                                    best_s)[2]
            best_s = \
            self.brute_force_params(answers, same, grain, best_l0, best_t,
                                    best_g, None)[3]
            print(best_l0, best_t, best_g, best_s)
        return best_l0, best_t, best_g, best_s
