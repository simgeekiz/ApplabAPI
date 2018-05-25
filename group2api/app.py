import numpy as np
from openpyxl import load_workbook
import datetime
from copy import deepcopy

from flask import Flask, request, jsonify, render_template, send_from_directory

app = Flask(__name__)

import os
import sys

sys.path.append('..')
import json
from pymongo import MongoClient
import pymongo

try:
    from .utils import calculations as calc
except ImportError:
    from utils import calculations as calc


# Connect to mongoLab
client = MongoClient(os.environ['MONGO_HOST'], int(os.environ['MONGO_PORT']))
db = client[os.environ['MONGO_DBNAME']]

# Authentication is only needed for db in mLab servers
# db.authenticate(os.environ['MONGO_DBUSER'], os.environ['MONGO_DBPASS'])

submissions = db['submissions']
users = db['users']


# helper function
def get_date_from_str(yyyymmdd):
    """ Return a datetime representation of the date
    :param yyyymmdd: str in format yyyy-mm-dd. division symbols can differ
    """
    split_token = yyyymmdd[4]
    year, month, day = [int(s) for s in yyyymmdd.split(split_token)]
    print("{}:{}:{}".format(year, month, day))
    return datetime.datetime(year, month, day, 0, 0, 0, 0)

@app.route('/')
def hello_world():
    """Display the home page

    Usage example : <hostname>/
    """
    return 'For documentation see /docs'


@app.route('/docs')
def serve_docs():
    """Display the documentations

    Usage example : <hostname>/docs
    """
    return render_template('index.html')


@app.route('/number_of_submissions')
def get_number_of_submissions():
    """Gives the total number of submissions

    Usage example : <hostname>/number_of_submissions

    Returns
    -------
    number_of_submissions : list of integers
        The list of user's ids
    """
    return jsonify(number_of_submissions=submissions.count())


@app.route('/users')
def get_user_list():
    """Lists the ids of the unique users

    Usage example : <hostname>/users

    Returns
    -------
    users : list of integers
        The list of user ids
    """
    user_list = submissions.distinct('UserId')
    return jsonify(user_list)


@app.route('/login_users')
def get_user_login_info():
    """Lists the userids, names and passwords

    Usage example : <hostname>/user_login_info

    Returns
    -------
    all_login_users : list of dictionaries
        The list of user login information
    """
    loginusers = users.find({}, {'_id': 0})
    all_login_users = [u for u in loginusers]
    return jsonify(all_login_users)


@app.route('/login_users/username=<username>')
def get_user_login_info_with_user(username):
    """Get the user id, name and password based on username

    Usage example : <hostname>/login_users/username=user15379
    
    Parameters
    ----------
    username : string
        Give user name of the user.

    Returns
    -------
    login_user : dictionary
        User login information
    """
    login_user_info = users.find({'UserName': username}, {'_id': 0})
    login_user = [u for u in login_user_info]
    return jsonify(login_user)


@app.route('/upload_login_users')
def upload_login_users():
    """Generates user information with default values

    Usage example : <hostname>/user_login_info
    """
    submissions_user_ids = submissions.distinct('UserId')
    users_user_ids = users.distinct('UserId')
    user_ids = [uid for uid in submissions_user_ids if
                uid not in users_user_ids]
    for uid in user_ids:
        user_info = {}
        user_info['UserId'] = uid
        user_info['UserName'] = 'user{}'.format(str(uid))
        user_info['Password'] = 1234
        users.insert(user_info)
    return jsonify(message="Users generated")


@app.route('/user/user_id=<user_id>')
def get_user_info(user_id):
    """Returns the list of activities belongs to a specific user

    Usage example : <hostname>/user/user_id=231412

    Parameters
    ----------
    user_id : integer or string
        Give the id of the user

    Returns
    -------
    user_info : list of dictionaries
        The list of user's activities
    """
    infos = submissions.find({'UserId': int(user_id)}, {'_id': 0})
    users_info = []
    for info in infos:
        users_info.append(info)
    return jsonify(users_info)


@app.route('/exercises/user_id=<user_id>&learning_obj_id=<learning_obj_id>')
def get_exercises(user_id, learning_obj_id):
    """Returns the list of exercises has been done by the user on a particlar learning objective

    Usage example : <hostname>/exercises/user_id=231412&learning_obj_id=554363

    Parameters
    ----------
    user_id : integer or string
        Give the id of the user

    learning_obj_id : integer or string
        Give the id of the learning objective

    Returns
    -------
    exercises : list of integers
        The list of unique exercise ids has been done by the user
    """
    exercises = submissions.find({'UserId': int(user_id),
                                  'LearningObjectiveId': int(
                                      learning_obj_id)}).distinct('ExerciseId')

    exercises_list = []
    for ex in exercises:
        if ex in ['', 'null ', 'null', 'NULL ', 'NULL', None]:
            exercises_list.append(None)
        else:
            exercises_list.append(int(ex))
    return jsonify(exercises_list)


@app.route(
    '/scores/user_id=<user_id>&learning_obj_id=<learning_obj_id>&exercise_id=<exercise_id>')
def get_scores(user_id, learning_obj_id, exercise_id):
    """Returns ability score of the user on an exercise for a learning objective

    Usage example : <hostname>/scores/user_id=231412&learning_obj_id=65745&exercise_id=342342

    Parameters
    ----------
    user_id : integer or string
        Give the id of the user

    learning_obj_id : integer or string
        Give the id of the learning objective

    exercise_id : integer or string
        Give the id of the exercise

    Returns
    -------
    scores : list of integers
        The list of the ability scores achieved by the user
    """
    scores = submissions.find(
        {'UserId': int(user_id),
         'LearningObjectiveId': int(learning_obj_id),
         'ExerciseId': int(exercise_id)},
        {'_id': 0, 'AbilityAfterAnswer': 1})

    scores_list = []
    for item in scores:
        score = item['AbilityAfterAnswer']
        if score in ['', 'null ', 'null', 'NULL ', 'NULL', None]:
            scores_list.append(None)
        else:
            scores_list.append(int(score))

    return jsonify(scores_list)


@app.route('/response_time_enrichment/user_id=<user_id>')
def response_time_enrichment(user_id='all'):
    """Calculates and adds the response time for each submission

    .. todo:: Run only on the data that doesn't have any response time

    Parameters
    ----------
    user_id : integer or string, (default='all')
        Give the id of the user. If given 'all', runs for
        each unique user in the database.
    """
    user_ids_list = submissions.distinct('UserId') if user_id == 'all' else [
        user_id]
    for user_id in user_ids_list:
        user_subs = submissions.find({'UserId': int(user_id)},
                                     {'_id': 1, 'SubmitDateTime': 1})
        user_subs = [sub for sub in user_subs]
        resp_time_enrich = calc.get_response_time(user_subs)
        for item in resp_time_enrich:
            submissions.update({'_id': item['_id']}, {
                '$set': {'ResponseTime': item['ResponseTime']}})
    return jsonify(message="Response times succesfully added.")


@app.route('/calculate_scores/user_id=<user_id>')
def calculate_scores(user_id='all'):
    """Calculates the ability scores for the given user or for all
    
    .. warning:: If run for all users, it takes a while
    
    Parameters
    ----------
    user_id : integer or string, (default='all')
        Give the id of the user. If given 'all', runs for
        each unique user in the database.
        
    Returns
    -------
    response : dictionary
        Contains a success message and a list of failed user ids.
    """
    response = {
        "message": "Calculated Ability scores succesfully added.",
        "failed_ids": []
    }
    user_ids_list = submissions.distinct('UserId') if user_id == 'all' else [user_id]
    for user_id in user_ids_list:
        user_subs = submissions.find({'UserId': int(user_id)},
                                     {'_id': 1, 'SubmitDateTime': 1,
                                      'ExerciseId': 1, 'ResponseTime': 1,
                                      'Correct': 1})
        user_subs = [sub for sub in user_subs]
        try:
            calculated_ability_score = calc.calculate_scores(user_subs)
        except KeyError:
            response['failed_ids'].append(user_id)
        else:
            for item in calculated_ability_score:
                submissions.update({'_id':item['_id']}, {'$set': {'AbilityScore': item['AbilityScore']}})
    return jsonify(response)

@app.route('/insert', methods=['POST'])
def insert():
    """Accepts data in JSON format and saves it to the file.

    In order to use it, the data should be *POSTed*.
    Data should consist of list of dictionaries.
    Usage example with Python requests library::

        import requests
        data = [{"SubmitDateTime":"2012-12-21 12:12:12.120",
                 "UserId":2464375364,
                 "ExerciseId":141464536,
                 "LearningObjectiveId":11424,
                 "Correct":1,
                 "AbilityAfterAnswer":14143}]
        requests.post("<hostname>/insert", json=data)
    """
    data_ = request.get_json()
    if type(data_) is not list:
        return jsonify(error="Data should be a list format.")
    try:
        for item in data_:
            submissions.insert(item)
    except:
        return jsonify(error="Error during insertion to database")
    
    try:
        response_time_enrichment(user_id='all')
    except:
        return jsonify(error="Error during response time enrichment")
    
    try:
        calculate_scores(user_id='all')
    except: 
        return jsonify(error="Error during calculate scores")
    
    try:
        upload_login_users()
    except: 
        return jsonify(error="Error during uploading users to database")
        
    return jsonify(message="Data succesfully inserted and saved.")


@app.route('/user_login', methods=['GET', 'POST'])
def user_login():
    """Checks if the requested used_id is present in the submission list,
    and if the password matches the one set by the user.
    """
    import pandas as pd
    # received = request.data.decode('utf-8')
    # TODO delete this when post request works
    received = '874250:1234'
    user_id, password = received.split(':')
    usr = "user" + user_id

    data_pd = pd.DataFrame(list(users.find()))
    
    try:
        with open('received.txt', 'w') as f:
            f.write(received)
            
        unames = data_pd['UserName'].tolist()
        if user_id == None or usr not in unames:
            return "Wrong username"

        elif usr in unames:
            ind = unames.index(usr)
            if password == str(data_pd["Password"][ind]):
                return "Correct"
            else:
                return "Wrong password"
    except:
        with open('received.txt', 'a') as f:
            f.write("failed")
        return "Error"

@app.route('/check_connection/')
def check():
    return "Hmm"

@app.route('/see_change')
def testing():
    return jsonify(message="this really is working :)")


@app.route('/calculate_m2m_coordinates/user_id=<user_id>')
@app.route('/calculate_m2m_coordinates/user_id=<user_id>/date=<date>')
def calculate_m2m_coordinates(user_id, date=None):
    """ Calculates coordinates that represent the learning of the
    moment-by-moment peaks.
    :param user_id: The user for which the coordinates are calculated
    :param date: optional, pass in yyyy-mm-dd format. Will result in getting
    just the coordinates for that day
    :return: The coordinates for which the
    """
    if date is None:
        coords = handler.get_coordinates_for_today(int(user_id))
    else:
        date = get_date_from_str(date)
        coords = handler.get_coordinates_for_today(int(user_id), date_id=date)
    return json.dumps(coords, cls=MyEncoder)


class MyEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        else:
            return super(MyEncoder, self).default(obj)


class DataHandler:
    def __init__(self,
                 fname="C:/Users/HP/Documents/study - offline/resultaten-radboud-all_anoniem.xlsx",
                 db=None):
        submissions = db
        self.dates = submissions.distinct('SubmitDateTime')
        self.user_ids = submissions.distinct('UserId')
        self.exercise_ids = submissions.distinct('ExerciseId')
        self.learn_obj_ids = submissions.distinct('LearningObjectiveId')
        self.corrects = submissions.distinct('Correct')
        self.ability_scores = submissions.distinct('AbilityAfterAnswer')
        self.m2m = MomentByMoment(self.user_ids, self.corrects, self)

    def get_max_row(self, column=1):
        val = self.ws.cell(row=1, column=column).value
        i = 1
        while val:
            i += 1
            val = self.ws.cell(row=i, column=column).value
        return i

    def get_color_row(self, column=1):
        for i in range(1, 20):
            color = self.ws.cell(row=i, column=column).fill.start_color.index
            if color != '00000000':
                return i

    def get_column(self, cid, start_row=2, end_row=None):
        if end_row is None:
            end_row = self.max_row
        if end_row <= start_row:
            end_row = start_row + 1
        column = []
        for i in range(start_row, end_row):
            value = self.ws.cell(row=i, column=cid).value
            if value or value == 0:
                column.append(value)
        return np.array(column)

    def get_users(self):
        return np.unique(self.user_ids)

    def get_corrects_from_user(self, user_id=0, method='all'):
        if method == 'all':
            return self.corrects[np.where(self.user_ids == user_id)]

    def get_dummy_graph_variables(self, user_id=0, method='all', oid=None):
        return self.m2m.get_p_T(user_id, method, oid)

    def get_graph_variables(self, user_id=0, method='all', oid=None):
        p_j = self.m2m.get_p_j(user_id=user_id, method=method,
                               objective_id=oid)
        self.graph_length = len(p_j)
        self.boundary_list, self.color_list = self.m2m.get_color_bars()
        return self.m2m.get_p_j(user_id=user_id, method=method,
                                objective_id=oid)

    def get_color_ids(self, fname='res/ID exercises.xlsx'):
        wb = load_workbook(fname)
        self.ws = wb.active
        pre = list(self.get_column(1, 3))
        c_bound = [self.get_color_row(i) for i in range(2, 5)]
        cin = list(self.get_column(2, 3, c_bound[0])) + \
              list(self.get_column(3, 3, c_bound[0])) + \
              list(self.get_column(4, 3, c_bound[0]))
        cex = list(self.get_column(2, c_bound[0])) + \
              list(self.get_column(3, c_bound[0])) + \
              list(self.get_column(4, c_bound[0]))
        aex = None
        raex = None
        post = list(self.get_column(5, 3))
        return (pre, cin, cex, aex, raex, post)

    def get_coordinates_for_today(self, user_id=0,
                                  date_id=None):
        # todo: change from answers to coordinates
        if date_id is None:
            answers = self.corrects[np.where(self.user_ids == user_id)]
        else:
            answers = self.corrects[np.where((self.user_ids == user_id) &
                                             (self.dates == date_id))]
        same = list(np.ones(len(answers)))
        try:
            same[0] = 0
        except IndexError:
            return []
        try:
            with open("../../data/parameters.csv") as f:
                parameters = np.genfromtxt(f)
                users = parameters[:1]
                parameters = parameters[np.where(str(user_id) == users)]
                parameters = parameters[1:]
        except FileNotFoundError:
            parameters = ParameterExtractor().smart_ssr(answers, same, 100, 3)
        user_m2m = deepcopy(self.m2m)
        user_m2m.set_initial_probabilities(*parameters)
        p_j = user_m2m.get_p_j(user_id, answers=answers)
        coordinates = self.from_p_j_to_coord(p_j)
        return coordinates

    def from_p_j_to_coord(self, p_j):
        summed_array = np.cumsum(p_j)[::-1]
        print(summed_array)
        direction = 1
        coords = []
        current = 0
        for speed in summed_array:
            current += direction * speed
            if abs(current) > speed:
                current = min(speed, max(-1 * speed, current))
                direction *= -1
                current += direction * speed
            coords.append(round(current, 2))
        max_val = coords[np.argmax(coords)]
        if max_val > 0:
            coords = [i/max_val for i in coords]
        return coords


class MomentByMoment:
    def __init__(self, user_ids, corrects, handler: DataHandler):
        self.p_l0 = 0.064
        self.p_T = 0.095
        self.p_G = 0.299
        self.p_S = 0.1
        self.p_ln = []
        self.user_ids = user_ids
        self.users = np.unique(user_ids)
        self.answers = corrects
        self.handler = handler
        self.list_p_T = []

    def set_initial_probabilities(self, l0, T, G, S):
        self.p_l0 = l0
        self.p_T = T
        self.p_G = G
        self.p_S = S

    def get_p_T(self, user_id, method='all', objective_id=None):
        return self.list_p_T[:]

    def get_p_j(self, user_id, method='all', objective_id=None, answers=None):
        user_answers = self.filter_answers(user_id, method, objective_id)
        if answers is not None:
            user_answers = answers
        p_ln = self.calculate_ln(user_answers)
        p_not_ln_t = [(1 - ln) * self.p_T for ln in p_ln]
        p_not_ln_not_t = [(1 - ln) * (1 - self.p_T) for ln in p_ln]
        self.list_p_T = p_ln[:-2]
        return self.calculate_p_j(user_answers, p_ln, p_not_ln_t,
                                  p_not_ln_not_t)

    def filter_answers(self, user_id, method, objectives_id):
        user_answers = self.answers[:]
        user_objectives = self.handler.learn_obj_ids[:]
        user_ids = self.user_ids[:]
        user_excs = self.handler.exercise_ids[:]
        user_dates = self.handler.dates[:]
        self.chosen_ids = np.where(
            (self.handler.learn_obj_ids == objectives_id) &
            (self.user_ids == user_id))

        user_answers = user_answers[self.chosen_ids]
        user_objectives = user_objectives[self.chosen_ids]
        user_excs = user_excs[self.chosen_ids]
        user_ids = user_ids[self.chosen_ids]
        if method not in ['all', 'first', 'second', 'all but first', 'last']:
            raise NotImplementedError
        if method == 'first':
            user_answers = self.filter_all_but_first(user_answers, user_excs)
        if method == 'second':
            user_answers = self.filter_all_but_second(user_answers, user_excs)
        if method == 'all but first':
            user_answers = self.filter_first(user_answers, user_excs)
        if method == 'last':
            user_answers = self.filter_all_but_last(user_answers, user_excs)
        self.excs = user_excs
        return user_answers

    def filter_all_but_first(self, answers, exercise_ids):
        exercises_processed = []
        return_answers = []
        for i in range(len(answers)):
            if exercise_ids[i] not in exercises_processed:
                return_answers.append(answers[i])
                exercises_processed.append(exercise_ids[i])
        return return_answers

    def filter_all_but_second(self, answers, exercise_ids):
        exercises_processed = []
        exercises_processed_twice = []
        return_answers = []
        for i in range(len(answers)):
            if exercise_ids[i] not in exercises_processed:
                exercises_processed.append(exercise_ids[i])
            else:
                if exercise_ids[i] not in exercises_processed_twice:
                    return_answers.append(answers[i])
                    exercises_processed_twice.append(exercise_ids[i])
        return return_answers

    def filter_first(self, answers, exercise_ids):
        exercises_processed = []
        return_answers = []
        for i in range(len(answers)):
            if exercise_ids[i] not in exercises_processed:
                exercises_processed.append(exercise_ids[i])
            else:
                return_answers.append(answers[i])
        return return_answers

    def filter_all_but_last(self, answers, exercise_ids):
        answers = answers[::-1]
        exercises_processed = []
        return_answers = []
        for i in range(len(answers)):
            if exercise_ids[i] not in exercises_processed:
                return_answers.append(answers[i])
                exercises_processed.append(exercise_ids[i])
        return return_answers[::-1]

    def calculate_ln(self, answers):
        p_ln = []
        for answer_id in range(len(answers)):
            if len(p_ln) == 0:
                k = self.p_l0
            else:
                k = p_ln[-1]
            s = self.p_S
            g = self.p_G
            if answers[answer_id] == 1:
                ln_prev_given_res = (k * (1 - s)) / (
                (k * (1 - s)) + ((1 - k) * g))
            else:
                ln_prev_given_res = (k * s) / ((k * s) + ((1 - k) * (1 - g)))
            p_ln.append(ln_prev_given_res + (1 - ln_prev_given_res) * self.p_T)
        return p_ln

    def calculate_p_j(self, answers, ln, n_ln_t, n_ln_n_t):
        p_j = []
        for a_id in range(len(answers) - 2):
            p_l = ln[a_id]
            p_nl_t = n_ln_t[a_id]
            x = n_ln_n_t[a_id]
            g = self.p_G
            t = self.p_T
            s = self.p_S
            if answers[a_id + 1] == 1:
                if answers[a_id + 2] == 1:  # RR
                    a_ln = (1 - s) ** 2
                    a_n_ln_n_t = g * (1 - t) * g + g * t * (1 - s)
                else:  # RW
                    a_ln = s * (1 - self.p_S)
                    a_n_ln_n_t = g * (1 - t) * (1 - g) + g * t * s
            else:
                if answers[a_id + 2] == 1:  # WR
                    a_ln = s * (1 - self.p_S)
                    a_n_ln_n_t = g * (1 - t) * (1 - g) + (1 - g) * t * (1 - s)
                else:  # WW
                    a_ln = s ** 2
                    a_n_ln_n_t = (1 - g) * (1 - t) * (1 - g) + (1 - g) * t * s
            a_n_ln_t = a_ln
            a12 = p_l * a_ln + p_nl_t * a_n_ln_t + x * a_n_ln_n_t
            p_j.append(a_n_ln_t * p_nl_t / a12)
        return p_j

    def get_color_bars(self):
        excs = self.excs
        bounds = [0]
        colors = ['blue', 'red', 'green', 'orange', 'purple', 'magenta']
        # find pre bound
        # print('finding pre-test')
        bound = 0
        e = excs[0]
        print(e)
        while e in self.handler.pre_ids:
            bound += 1
            e = excs[bound]
        # print(e)
        bounds.append(bound)

        # Find class instruction
        # print('finding instruction exercises')
        bound += 1
        e = excs[bound + 1]
        while e in self.handler.c_in_ids:
            bound += 1
            e = excs[bound]
        # print(e)
        bounds.append(bound)

        # Find class exercise
        # print('finding class exercises')
        bound += 1
        e = excs[bound + 1]
        while e in self.handler.c_ex_ids:
            bound += 1
            e = excs[bound]
        # print(e)
        bounds.append(bound)

        # Find class adaptive
        bound += 1
        dates = self.handler.dates[self.chosen_ids]
        d = dates[bound]
        while dates[bound + 1] == d:
            bound += 1
        # print(excs[bound])
        bounds.append(bound)

        # Find repetition adaptive
        # print('finding repeated adaptive exercises')
        bound += 1
        e = excs[bound + 1]
        while e not in self.handler.post_ids:
            bound += 1
            e = excs[bound]
        # print(e)
        bounds.append(bound)

        # print('find post-test')
        while bound < len(excs) - 1:
            bound += 1
        # print(excs[bound])
        bounds.append(bound)
        # print(bounds)
        return (bounds, colors)


class ParameterExtractor:
    def __init__(self):
        # params = [L0, T, G, S]
        self.params_min = [1e-15 for i in range(4)]
        self.params_max = [1.0, 1.0, 0.3, 0.1]

    def brute_force_params(self, answers, same, grain=100, L0_fix=None,
                           T_fix=None, G_fix=None, S_fix=None, ):
        # set ranges up
        best_l0 = L0_fix
        best_t = T_fix
        best_g = G_fix
        best_s = S_fix
        best_SSR = len(answers) * 999999999999991
        L0_range = self.get_range(L0_fix, 0, grain)
        T_range = self.get_range(T_fix, 1, grain)
        G_range = self.get_range(G_fix, 2, grain)
        S_range = self.get_range(S_fix, 3, grain)
        for L0 in L0_range:
            # print('------------------------------------\nL0 is now at:{}'.format(L0))
            for T in T_range:
                for G in G_range:
                    for S in S_range:
                        new_SSR = self.get_s_s_r(L0, T, G, S, answers, same)
                        if new_SSR < best_SSR:
                            best_l0, best_t, best_g, best_s, best_SSR = [L0, T,
                                                                         G, S,
                                                                         new_SSR]
                        # print('best parameters now at L0:{}, T:{}, G:{}, S:{}'.format(L0,
                        # 	T, G, S))
        return best_l0, best_t, best_g, best_s

    def get_s_s_r(self, L0, T, G, S, answers, sames=None):
        SSR = 0.0
        S = max(1E-15, S)
        T = max(1E-15, T)
        G = max(1E-15, G)
        L0 = max(1E-15, L0)
        L = L0
        for same, answer in zip(sames, answers):
            if same == 0:
                L = L0
            # print(L, T, G, S)
            SSR += (answer - (L * (1.0 - S) + (1.0 - L) * G)) ** 2
            if answer == 0:
                L_given_answer = (L * S) / ((L * S) + ((1.0 - L) * (1.0 - G)))
            else:
                L_given_answer = (L * (1.0 - S)) / (
                (L * (1.0 - S)) + ((1.0 - L) * G))
            if not L_given_answer:
                print('huh')
            L = L_given_answer + (1.0 - L_given_answer) * T
        return SSR

    def get_range(self, possible_range, par_id, grain):
        if possible_range is None:
            return np.linspace(self.params_min[par_id],
                               self.params_max[par_id],
                               int(grain * self.params_max[par_id]),
                               endpoint=False)[1:]
        return [possible_range]

    def smart_ssr(self, answers, same, grain, iterations):
        best_l0 = \
        self.brute_force_params(answers, same, grain, None, 0.0, 0.0, 0.0)[0]
        best_t = \
        self.brute_force_params(answers, same, grain, 0.0, None, 0.0, 0.0)[1]
        best_g = \
        self.brute_force_params(answers, same, grain, 0.0, 0.0, None, 0.0)[2]
        best_s = \
        self.brute_force_params(answers, same, grain, 0.0, 0.0, 0.0, None)[3]
        for i in range(iterations):
            print("best is {}".format([best_l0, best_t, best_g, best_s]))
            best_l0 = \
            self.brute_force_params(answers, same, grain, None, best_t, best_g,
                                    best_s)[0]
            best_t = \
            self.brute_force_params(answers, same, grain, best_l0, None,
                                    best_g,
                                    best_s)[1]
            best_g = \
            self.brute_force_params(answers, same, grain, best_l0, best_t,
                                    None,
                                    best_s)[2]
            best_s = \
            self.brute_force_params(answers, same, grain, best_l0, best_t,
                                    best_g, None)[3]
        return best_l0, best_t, best_g, best_s

handler = None

if __name__ == '__main__':
    handler = DataHandler('', submissions)
    app.run()
